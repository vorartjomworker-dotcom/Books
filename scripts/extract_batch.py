#!/usr/bin/env python3
"""Extract one exact block range from an ML4T XLSX registry into CSV."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

ID_RE = re.compile(r"^(CH\d{2})-B(\d{4})$")


def read_registry(path: Path, sheet: str = "Блоки") -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook[sheet]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(rows)]
    required = {"chapter_id", "block_id", "sequence_number", "english_text", "russian_text", "source_html", "translation_status"}
    missing = required - set(header)
    if missing:
        raise ValueError(f"missing registry columns: {sorted(missing)}")
    records = [dict(zip(header, row)) for row in rows if any(value is not None for value in row)]
    return header, records


def select_range(records: list[dict[str, object]], chapter: str, start: int, end: int) -> list[dict[str, object]]:
    selected = [record for record in records if record.get("chapter_id") == chapter and start <= int(record.get("sequence_number") or -1) <= end]
    expected = list(range(start, end + 1))
    actual = [int(record["sequence_number"]) for record in selected]
    if actual != expected:
        raise ValueError(f"range mismatch: expected {expected[0]}..{expected[-1]}, got {actual}")
    for record in selected:
        match = ID_RE.fullmatch(str(record["block_id"]))
        if not match or match.group(1) != chapter or int(match.group(2)) != int(record["sequence_number"]):
            raise ValueError(f"identifier mismatch: {record['block_id']}")
        if record.get("translation_status") in {"QA_PASS", "ACCEPTED", "READY_FOR_PUBLICATION"}:
            raise ValueError(f"accepted block in requested range: {record['block_id']}")
    return selected


def write_csv(path: Path, header: list[str], records: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("registry", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--chapter", required=True)
    parser.add_argument("--start", required=True, type=int)
    parser.add_argument("--end", required=True, type=int)
    args = parser.parse_args()
    header, records = read_registry(args.registry)
    selected = select_range(records, args.chapter, args.start, args.end)
    write_csv(args.output, header, selected)
    print(f"Extracted {len(selected)} blocks: {selected[0]['block_id']}..{selected[-1]['block_id']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

