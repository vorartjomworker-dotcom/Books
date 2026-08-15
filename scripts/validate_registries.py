#!/usr/bin/env python3
"""Validate ML4T chapter registries, QA issue registers, and snapshot hashes."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


BLOCK_RE = re.compile(r"^(CH\d{2})-B(\d{4})$")
ERROR_RE = re.compile(r"^CH\d{2}-ERR-\d{3,}$")

ALIASES = {
    "block_id": {"block id", "block_id", "id блока", "идентификатор блока"},
    "batch_id": {"batch id", "batch_id", "id партии", "идентификатор партии"},
    "status": {"status", "статус", "qa status", "qa_status", "итоговый статус"},
    "english": {
        "english",
        "english text",
        "english_text",
        "source text",
        "source_text",
        "оригинал",
        "исходный текст",
    },
    "russian": {
        "russian",
        "russian text",
        "russian_text",
        "target text",
        "target_text",
        "translation",
        "перевод",
        "русский текст",
    },
    "error_id": {"error id", "error_id", "id ошибки", "идентификатор ошибки"},
}


def normalize(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\ufeff", "").strip().lower().replace("-", " ")
    return re.sub(r"\s+", " ", text)


def column_map(row: Sequence[object]) -> dict[str, int]:
    normalized = [normalize(value) for value in row]
    result: dict[str, int] = {}
    for logical_name, aliases in ALIASES.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                result[logical_name] = index
                break
    return result


def find_header(rows: Iterable[Sequence[object]], required: str) -> tuple[int, dict[str, int], list[object]] | None:
    for row_number, row in enumerate(rows, start=1):
        values = list(row)
        columns = column_map(values)
        if required in columns:
            return row_number, columns, values
        if row_number >= 25:
            break
    return None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class Report:
    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.passed: list[str] = []

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def ok(self, message: str) -> None:
        self.passed.append(message)


def workbook_rows(path: Path, required_header: str):
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            probe = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True))
            found = find_header(probe, required_header)
            if found:
                header_row, columns, header = found
                rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))
                return sheet.title, header_row, columns, header, rows
    finally:
        workbook.close()
    return None


def validate_chapter(root: Path, spec: dict, allowed_statuses: set[str], translated_statuses: set[str], report: Report) -> None:
    chapter = spec["chapter"]
    path = root / spec["path"]
    if not path.is_file():
        report.error(f"{chapter}: отсутствует {spec['path']}")
        return
    try:
        table = workbook_rows(path, "block_id")
    except Exception as exc:  # corrupt or unsupported workbook
        report.error(f"{chapter}: файл XLSX не читается: {exc}")
        return
    if table is None:
        report.error(f"{chapter}: не найден столбец block_id в первых 25 строках ни одного листа")
        return

    sheet, header_row, columns, _header, rows = table
    records: list[tuple[str, Sequence[object]]] = []
    for row in rows:
        value = row[columns["block_id"]] if columns["block_id"] < len(row) else None
        block_id = "" if value is None else str(value).strip()
        if BLOCK_RE.fullmatch(block_id):
            records.append((block_id, row))

    expected_count = int(spec["expected_count"])
    ids = [block_id for block_id, _ in records]
    duplicates = sorted(value for value, count in Counter(ids).items() if count > 1)
    if duplicates:
        report.error(f"{chapter}: повторяющиеся block_id: {', '.join(duplicates[:10])}")

    foreign = sorted(value for value in set(ids) if not value.startswith(chapter + "-B"))
    if foreign:
        report.error(f"{chapter}: идентификаторы другой главы: {', '.join(foreign[:10])}")

    actual_numbers = sorted(int(BLOCK_RE.fullmatch(value).group(2)) for value in set(ids))
    expected_numbers = list(range(1, expected_count + 1))
    if actual_numbers != expected_numbers:
        missing = sorted(set(expected_numbers) - set(actual_numbers))
        extra = sorted(set(actual_numbers) - set(expected_numbers))
        if missing:
            report.error(f"{chapter}: пропущены номера блоков: {missing[:20]}")
        if extra:
            report.error(f"{chapter}: номера за пределами диапазона: {extra[:20]}")
    if len(ids) != expected_count:
        report.error(f"{chapter}: найдено {len(ids)} строк блоков, ожидалось {expected_count}")

    status_index = columns.get("status")
    english_index = columns.get("english")
    russian_index = columns.get("russian")
    unknown_statuses: set[str] = set()
    missing_translation: list[str] = []
    non_passed_prefix: list[str] = []
    qa_pass_through = int(spec.get("qa_pass_through", 0))
    for block_id, row in records:
        status = ""
        if status_index is not None and status_index < len(row):
            status = "" if row[status_index] is None else str(row[status_index]).strip().upper()
            if status and status not in allowed_statuses:
                unknown_statuses.add(status)
            match = BLOCK_RE.fullmatch(block_id)
            if match and int(match.group(2)) <= qa_pass_through and status != "QA_PASS":
                non_passed_prefix.append(block_id)
        if status in translated_statuses and english_index is not None and russian_index is not None:
            english = row[english_index] if english_index < len(row) else None
            russian = row[russian_index] if russian_index < len(row) else None
            if not normalize(english) or not normalize(russian):
                missing_translation.append(block_id)

    if unknown_statuses:
        report.error(f"{chapter}: недопустимые статусы: {', '.join(sorted(unknown_statuses))}")
    if non_passed_prefix:
        report.error(
            f"{chapter}: контрольный диапазон 0001–{qa_pass_through:04d} не имеет QA_PASS: "
            f"{', '.join(non_passed_prefix[:20])}"
        )
    if missing_translation:
        report.error(f"{chapter}: пустой исходный текст или перевод при финальном статусе: {', '.join(missing_translation[:20])}")
    if status_index is None:
        report.warn(f"{chapter}: столбец статуса не распознан; проверка статусов пропущена")
    if english_index is None or russian_index is None:
        report.warn(f"{chapter}: пара столбцов EN/RU не распознана; проверка заполненности перевода пропущена")

    if not any(message.startswith(chapter + ":") for message in report.errors):
        report.ok(f"{chapter}: {len(ids)} уникальных блоков, диапазон 0001–{expected_count:04d}, лист «{sheet}», заголовок {header_row}")


def validate_issue_registry(path: Path, report: Report) -> None:
    relative = path.as_posix()
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
            found = find_header(rows[:25], "error_id")
            if not found:
                report.error(f"{relative}: не найден столбец error_id")
                return
            header_row, columns, _ = found
            data = rows[header_row:]
        else:
            table = workbook_rows(path, "error_id")
            if table is None:
                report.warn(f"{relative}: структурированный столбец error_id не найден; проверена только читаемость XLSX")
                return
            _sheet, _header_row, columns, _header, data = table
        index = columns["error_id"]
        ids = [str(row[index]).strip() for row in data if index < len(row) and row[index] not in (None, "")]
        duplicates = sorted(value for value, count in Counter(ids).items() if count > 1)
        malformed = sorted(value for value in ids if not ERROR_RE.fullmatch(value))
        if duplicates:
            report.error(f"{relative}: повторяющиеся error_id: {', '.join(duplicates[:10])}")
        if malformed:
            report.error(f"{relative}: неверный формат error_id: {', '.join(malformed[:10])}")
        if not duplicates and not malformed:
            report.ok(f"{relative}: {len(ids)} уникальных записей ошибок")
    except Exception as exc:
        report.error(f"{relative}: реестр ошибок не читается: {exc}")


def validate_checksums(root: Path, manifest_name: str, report: Report) -> None:
    manifest = root / manifest_name
    if not manifest.is_file():
        report.error(f"Отсутствует файл контрольных сумм {manifest_name}")
        return
    checked = 0
    for line_number, raw in enumerate(manifest.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split(None, 1)
        if len(parts) != 2 or not re.fullmatch(r"[0-9a-fA-F]{64}", parts[0]):
            report.error(f"{manifest_name}:{line_number}: неверный формат строки")
            continue
        expected, file_name = parts[0].lower(), parts[1].strip()
        path = root / file_name
        if not path.is_file():
            report.error(f"{manifest_name}:{line_number}: отсутствует {file_name}")
            continue
        actual = sha256(path)
        if actual != expected:
            report.error(f"{file_name}: SHA-256 {actual}, ожидалось {expected}")
        else:
            checked += 1
    if checked:
        report.ok(f"Контрольные суммы: подтверждено файлов — {checked}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--config", type=Path, default=Path("config/registry_expectations.json"))
    parser.add_argument("--strict-warnings", action="store_true")
    args = parser.parse_args()
    root = args.root.resolve()
    config_path = args.config if args.config.is_absolute() else root / args.config
    config = json.loads(config_path.read_text(encoding="utf-8"))
    report = Report()

    allowed_statuses = {str(value).upper() for value in config["allowed_statuses"]}
    translated_statuses = {str(value).upper() for value in config["translated_statuses"]}
    for chapter in config["chapters"]:
        validate_chapter(root, chapter, allowed_statuses, translated_statuses, report)

    for pattern in config["issue_registry_globs"]:
        matches = sorted(root.glob(pattern))
        if not matches:
            report.warn(f"Не найден ни один реестр ошибок по шаблону {pattern}")
        for path in matches:
            validate_issue_registry(path, report)

    for pattern in config["forbidden_source_globs"]:
        for path in root.glob(pattern):
            if path.is_file():
                report.error(f"Запрещённый исходный файл отслеживается в Git: {path.relative_to(root)}")

    for manifest in config["checksum_manifests"]:
        validate_checksums(root, manifest, report)

    print("ML4T registry validation")
    for message in report.passed:
        print(f"PASS: {message}")
    for message in report.warnings:
        print(f"WARN: {message}")
    for message in report.errors:
        print(f"FAIL: {message}")
    print(f"Summary: {len(report.passed)} passed, {len(report.warnings)} warnings, {len(report.errors)} errors")
    return 1 if report.errors or (args.strict_warnings and report.warnings) else 0


if __name__ == "__main__":
    sys.exit(main())
